"""
Management command para importar alunos e turmas a partir do Excel SEED-PR.

Uso:
    python manage.py importar_excel "Alunos SEED-PR - Toni Coimbra.xlsx"
    python manage.py importar_excel "Alunos SEED-PR - Toni Coimbra.xlsx" --dry-run
"""

from django.core.management.base import BaseCommand
from django.db import transaction

from alunos.models import Aluno
from turmas.models import Matricula, Turma

# Mapeamento: nome da aba -> (nome da turma, codigo, periodo, ano_letivo)
SHEET_MAP = {
    # --- Turno Manhã (1º A) ---
    "Análise e Método para Sistemas ": (
        "Análise e Método para Sistemas - 1º Ano A Manhã",
        "AMS-1A-M-2026",
        "1º Ano A Manhã",
        2026,
    ),
    "Introdução à Computação - 1º An": (
        "Introdução à Computação - 1º Ano A Manhã",
        "IC-1A-M-2026",
        "1º Ano A Manhã",
        2026,
    ),
    "Programação Front-End - 2º Ano ": (
        "Programação Front-End - 2º Ano A Manhã",
        "PFE-2A-M-2026",
        "2º Ano A Manhã",
        2026,
    ),
    "Inovação Tec e Empreend - 2º An": (
        "Inovação Tecnológica e Empreendedorismo - 2º Ano A Manhã",
        "ITE-2A-M-2026",
        "2º Ano A Manhã",
        2026,
    ),
    "Programação de Sistemas - 3º An": (
        "Programação de Sistemas - 3º Ano A Manhã",
        "PS-3A-M-2026",
        "3º Ano A Manhã",
        2026,
    ),
    "Análise e Projeto de Sistemas -": (
        "Análise e Projeto de Sistemas - 3º Ano A Manhã",
        "APS-3A-M-2026",
        "3º Ano A Manhã",
        2026,
    ),
    # --- Turno Noite (3º C) ---
    "Página1": (
        "Programação de Sistemas - 3º Ano C Noite",
        "PS-3C-N-2026",
        "3º Ano C Noite",
        2026,
    ),
    "Página2": (
        "Introdução à Computação - 1º Ano C Noite",
        "IC-1C-N-2026",
        "1º Ano C Noite",
        2026,
    ),
    "Página3": (
        "Inovação Tecnológica e Empreendedorismo - 2º Ano C Noite",
        "ITE-2C-N-2026",
        "2º Ano C Noite",
        2026,
    ),
    "Página4": (
        "Análise e Projeto de Sistemas - 3º Ano C Noite",
        "APS-3C-N-2026",
        "3º Ano C Noite",
        2026,
    ),
}


class Command(BaseCommand):
    help = "Importa alunos e turmas do arquivo Excel SEED-PR"

    def add_arguments(self, parser):
        parser.add_argument("arquivo", type=str, help="Caminho do arquivo .xlsx")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Apenas mostra o que seria feito, sem alterar o banco",
        )
        parser.add_argument(
            "--incluir-paginas",
            type=str,
            nargs="*",
            help=(
                "Incluir abas Página com nome de turma. "
                "Formato: 'Página1:Nome da Turma:CODIGO:Periodo'"
            ),
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write(self.style.ERROR("Instale openpyxl: pip install openpyxl"))
            return

        arquivo = options["arquivo"]
        dry_run = options["dry_run"]

        wb = openpyxl.load_workbook(arquivo, read_only=True)

        # Processar mapeamento extra de --incluir-paginas
        sheet_map = dict(SHEET_MAP)
        if options.get("incluir_paginas"):
            for spec in options["incluir_paginas"]:
                parts = spec.split(":")
                if len(parts) < 4:
                    self.stderr.write(
                        self.style.ERROR(f"Formato invalido: {spec}. Use Aba:Nome:Codigo:Periodo")
                    )
                    return
                sheet_map[parts[0]] = (parts[1], parts[2], parts[3], 2026)

        stats = {
            "turmas_criadas": 0,
            "turmas_existentes": 0,
            "alunos_criados": 0,
            "alunos_existentes": 0,
            "matriculas_criadas": 0,
            "matriculas_existentes": 0,
        }

        with transaction.atomic():
            for sheet_name in wb.sheetnames:
                if sheet_name not in sheet_map:
                    self.stdout.write(
                        self.style.WARNING(f"Aba ignorada (sem mapeamento): {sheet_name}")
                    )
                    continue

                nome_turma, codigo, periodo, ano_letivo = sheet_map[sheet_name]
                self.stdout.write(f"\n{'='*60}")
                self.stdout.write(self.style.HTTP_INFO(f"Processando: {nome_turma} ({codigo})"))

                # Criar ou buscar turma
                turma, turma_criada = Turma.objects.get_or_create(
                    codigo=codigo,
                    defaults={
                        "nome": nome_turma,
                        "periodo": periodo,
                        "ano_letivo": ano_letivo,
                        "ativa": True,
                    },
                )
                if turma_criada:
                    stats["turmas_criadas"] += 1
                    self.stdout.write(self.style.SUCCESS(f"  Turma CRIADA: {turma}"))
                else:
                    stats["turmas_existentes"] += 1
                    self.stdout.write(f"  Turma existente: {turma}")

                # Processar alunos da aba
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    nome = row[1]
                    email = row[2]
                    if not nome or not email:
                        continue

                    nome = str(nome).strip()
                    email = str(email).strip().lower()

                    # Criar ou buscar aluno
                    aluno, aluno_criado = Aluno.objects.get_or_create(
                        email=email,
                        defaults={"nome": nome},
                    )
                    if aluno_criado:
                        stats["alunos_criados"] += 1
                    else:
                        stats["alunos_existentes"] += 1

                    # Criar matrícula
                    matricula, mat_criada = Matricula.objects.get_or_create(
                        aluno=aluno,
                        turma=turma,
                        defaults={"ativa": True},
                    )
                    if mat_criada:
                        stats["matriculas_criadas"] += 1
                        self.stdout.write(f"    + {nome} ({email})")
                    else:
                        stats["matriculas_existentes"] += 1
                        if not matricula.ativa:
                            matricula.ativa = True
                            matricula.save(update_fields=["ativa"])

            if dry_run:
                self.stdout.write(self.style.WARNING("\n[DRY-RUN] Revertendo todas as alteracoes..."))
                transaction.set_rollback(True)

        # Resumo
        self.stdout.write(f"\n{'='*60}")
        self.stdout.write(self.style.SUCCESS("RESUMO DA IMPORTACAO"))
        self.stdout.write(f"  Turmas criadas:      {stats['turmas_criadas']}")
        self.stdout.write(f"  Turmas existentes:   {stats['turmas_existentes']}")
        self.stdout.write(f"  Alunos criados:      {stats['alunos_criados']}")
        self.stdout.write(f"  Alunos existentes:   {stats['alunos_existentes']}")
        self.stdout.write(f"  Matriculas criadas:  {stats['matriculas_criadas']}")
        self.stdout.write(f"  Matriculas existentes: {stats['matriculas_existentes']}")

        if dry_run:
            self.stdout.write(self.style.WARNING("Nenhuma alteracao foi salva (--dry-run)"))
